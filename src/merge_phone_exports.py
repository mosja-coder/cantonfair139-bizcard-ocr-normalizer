#!/usr/bin/env python3
"""
Merge partition ocr_card_contacts.xlsx (phone_processed sheet) into one workbook.

Usage:
  python merge_phone_exports.py
  python merge_phone_exports.py --output "../data/汇总/广交会139届_名片手机号汇总.xlsx"
"""

from __future__ import annotations

import argparse
from pathlib import Path
from typing import Dict, List, Tuple

from openpyxl import Workbook, load_workbook

from partitions import DATASET_COL, PARTITIONS, SOURCE_COL
from paths import SOURCE_DATASET_LABEL, deliverables_dir, project_root, run_dir


def read_phone_sheet(xlsx_path: Path) -> Tuple[List[str], List[Dict[str, object]]]:
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb["phone_processed"]
    rows_iter = ws.iter_rows(values_only=True)
    headers = [str(h) if h is not None else "" for h in next(rows_iter)]
    records: List[Dict[str, object]] = []
    for row in rows_iter:
        if row is None or all(v is None or v == "" for v in row):
            continue
        records.append({headers[i]: row[i] for i in range(len(headers))})
    wb.close()
    return headers, records


def merge_all(root: Path) -> Tuple[List[str], List[Dict[str, object]], List[Tuple[str, int]]]:
    base_headers: List[str] = []
    merged: List[Dict[str, object]] = []
    summary: List[Tuple[str, int]] = []

    for run_name, source_label in PARTITIONS:
        xlsx = run_dir(root, run_name) / "ocr_card_contacts.xlsx"
        if not xlsx.exists():
            raise FileNotFoundError(f"缺少分区文件: {xlsx}")

        headers, records = read_phone_sheet(xlsx)
        if not base_headers:
            base_headers = headers

        for rec in records:
            row = dict(rec)
            row[SOURCE_COL] = source_label
            row[DATASET_COL] = SOURCE_DATASET_LABEL
            merged.append(row)
        summary.append((source_label, len(records)))

    out_headers = [SOURCE_COL, DATASET_COL] + [h for h in base_headers if h not in (SOURCE_COL, DATASET_COL)]
    return out_headers, merged, summary


def write_workbook(output_path: Path, headers: List[str], rows: List[Dict[str, object]], summary: List[Tuple[str, int]]) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()

    ws = wb.active
    ws.title = "phone_processed"
    ws.append(headers)
    for row in rows:
        ws.append([row.get(h, "") for h in headers])

    ws_sum = wb.create_sheet("汇总统计")
    ws_sum.append(["source_label", "row_count"])
    for label, count in summary:
        ws_sum.append([label, count])
    ws_sum.append(["合计", sum(c for _, c in summary)])

    wb.save(output_path)


def main() -> int:
    parser = argparse.ArgumentParser(description="汇总 12 个分区的 phone_processed 到单一 Excel")
    parser.add_argument(
        "--output",
        default="",
        help="输出路径，默认 data/deliverables/广交会139届_名片手机号汇总.xlsx",
    )
    args = parser.parse_args()

    root = project_root()
    output = (
        Path(args.output).expanduser().resolve()
        if args.output
        else deliverables_dir(root) / "广交会139届_名片手机号汇总.xlsx"
    )

    headers, merged, summary = merge_all(root)
    write_workbook(output, headers, merged, summary)

    print(f"已汇总 {len(summary)} 个分区，共 {len(merged)} 行")
    for label, count in summary:
        print(f"  {label}: {count}")
    print(f"输出: {output}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
