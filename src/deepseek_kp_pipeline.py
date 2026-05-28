#!/usr/bin/env python3
"""
DeepSeek post-processing on existing OCR card results (default model: deepseek-v4-flash).

Reads ocr_card_results.jsonl (no re-OCR), calls DeepSeek to standardize fields,
filter key persons (KP), normalize +86 mobiles, and export merged xlsx.

Usage:
  python deepseek_kp_pipeline.py --resume
  python deepseek_kp_pipeline.py --partition 2期-PartA --limit 20
"""

from __future__ import annotations

import argparse
import json
import os
import re
import sys
import threading
import time
import urllib.error
import urllib.request
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path
from typing import Any, Dict, List, Optional, Set, Tuple

from dotenv import load_dotenv
from openpyxl import Workbook, load_workbook

from checkpoint import atomic_write_json, utc_now_iso
from partitions import DATASET_COL, PARTITIONS, SOURCE_COL
from paths import (
    SOURCE_DATASET_LABEL,
    deepseek_cache_dir,
    deliverables_dir,
    project_root,
    run_dir,
)
from process_phone_sheet import normalize_phones

DEEPSEEK_API_URL = "https://api.deepseek.com/chat/completions"
DEFAULT_MODEL = "deepseek-v4-flash"
DEFAULT_BATCH_SIZE = 6
DEFAULT_WORKERS = 6
DEFAULT_HEARTBEAT_SEC = 15.0


def configure_stdio_unbuffered() -> None:
    """Ensure progress lines appear immediately in Cursor terminal."""
    os.environ.setdefault("PYTHONUNBUFFERED", "1")
    if hasattr(sys.stdout, "reconfigure"):
        try:
            sys.stdout.reconfigure(line_buffering=True)
        except Exception:
            pass


class ProgressReporter:
    """Thread-safe terminal progress + periodic heartbeat while API calls run."""

    def __init__(
        self,
        total: int,
        model: str,
        *,
        initial_done: int = 0,
        initial_tokens: int = 0,
        heartbeat_sec: float = DEFAULT_HEARTBEAT_SEC,
        enabled: bool = True,
    ) -> None:
        self.total = total
        self.model = model
        self.heartbeat_sec = heartbeat_sec
        self.enabled = enabled
        self.lock = threading.Lock()
        self.done = initial_done
        self.tokens = initial_tokens
        self.status = "running"
        self.completed_batches = 0
        self.total_batches = 0
        self.in_flight = 0
        self._stop = False
        self._thread: Optional[threading.Thread] = None
        self._started_at = time.time()
        self._initial_done = initial_done

    def start(self, total_batches: int) -> None:
        self.total_batches = total_batches
        if not self.enabled:
            return
        self._print_line(tag="启动")
        if self.heartbeat_sec > 0:
            self._thread = threading.Thread(target=self._heartbeat_loop, daemon=True)
            self._thread.start()

    def stop(self, final_status: str = "completed") -> None:
        with self.lock:
            self.status = final_status
            self._stop = True
        if self.enabled:
            self._print_line(tag="完成")
        if self._thread:
            self._thread.join(timeout=1.0)

    def batch_started(self) -> None:
        with self.lock:
            self.in_flight += 1

    def batch_finished(
        self,
        done: int,
        tokens: int,
        status: str,
        completed_batches: int,
    ) -> None:
        with self.lock:
            self.done = done
            self.tokens = tokens
            self.status = status
            self.completed_batches = completed_batches
            self.in_flight = max(0, self.in_flight - 1)
            if self.enabled:
                self._print_line(tag="进度")

    def _heartbeat_loop(self) -> None:
        while True:
            time.sleep(self.heartbeat_sec)
            with self.lock:
                if self._stop:
                    break
                if self.enabled:
                    self._print_line(tag="心跳")

    def _format_eta(self, remain: int) -> str:
        elapsed = max(time.time() - self._started_at, 1.0)
        session_done = self.done - self._initial_done
        rate = session_done / elapsed if session_done > 0 else 0.0
        if rate <= 0:
            return "ETA --"
        eta_sec = remain / rate
        if eta_sec < 3600:
            return f"ETA {int(eta_sec // 60)}m"
        return f"ETA {eta_sec / 3600:.1f}h"

    def _print_line(self, tag: str = "进度") -> None:
        remain = max(self.total - self.done, 0)
        pct = (self.done / self.total * 100) if self.total else 0.0
        ts = time.strftime("%H:%M:%S")
        eta = self._format_eta(remain)
        msg = (
            f"[{ts}] [{tag}] DeepSeek | "
            f"已完成 {self.done}/{self.total} ({pct:.1f}%) | "
            f"剩余 {remain} | "
            f"状态 {self.status} | "
            f"tokens {self.tokens:,} | "
            f"batch {self.completed_batches}/{self.total_batches} | "
            f"进行中 {self.in_flight} | "
            f"{eta} | "
            f"{self.model}"
        )
        print(msg, flush=True)


def resolve_model(explicit: str = "") -> str:
    return (explicit or os.getenv("DEEPSEEK_MODEL", DEFAULT_MODEL)).strip() or DEFAULT_MODEL


def model_pipeline_tag(model: str) -> str:
    return model.replace("-", "_").replace(".", "_")


def default_output_xlsx(root: Path, model: str) -> Path:
    short = "dsv4flash" if "flash" in model else "dsv4p" if "pro" in model else model.replace("-", "")
    return deliverables_dir(root) / f"广交会139届_名片手机号汇总_{short}.xlsx"

SYSTEM_PROMPT = """你是广交会名片数据结构化专家。用户会给你一批 OCR 识别出的名片文本行（可能有噪声、乱码、换行错误）。

请对每张名片输出结构化结果，要求：
1. **标准化** name（联系人姓名）、title（职位）、company（公司名）、address（地址）
   - 纠正明显 OCR 错误，保留有效英文/中英文混合信息
   - 剔除脏值：乱码、纯符号、无意义片段、邮箱/网址混入姓名等
   - 无法可靠识别则对应字段留空字符串
2. **mobiles**：仅提取中国大陆 11 位手机号（1 开头的 11 位数字），去重
3. **is_key_person**：判断是否为决策层/管理层 KP
   - true：董事长、总经理、总裁、副总、总监、经理、主管、负责人、创始人、Owner、CEO、GM、Managing Director、Director、Manager 等
   - false：普通销售、业务员、文员、前台、助理（无管理职级）等
4. **kp_role**：简短中文或英文职级标签（如「总经理」「Sales Manager」），非 KP 可留空

严格返回 JSON 对象，格式：
{"cards":[{"card_key":"...","name":"","title":"","company":"","address":"","mobiles":[],"is_key_person":false,"kp_role":""}]}
card_key 必须与输入完全一致。不要输出 markdown。"""


def bootstrap_dotenv(root: Path) -> None:
    load_dotenv(root / ".env", override=False)


def make_card_key(source_label: str, card: Dict[str, Any]) -> str:
    return f"{source_label}|{card.get('file')}|{card.get('page')}|{card.get('card_no')}"


def load_cards_from_partitions(root: Path, partition_filter: str = "") -> List[Dict[str, Any]]:
    cards: List[Dict[str, Any]] = []
    for run_name, source_label in PARTITIONS:
        if partition_filter and source_label != partition_filter:
            continue
        jsonl = run_dir(root, run_name) / "ocr_card_results.jsonl"
        if not jsonl.exists():
            raise FileNotFoundError(f"缺少 OCR 结果: {jsonl}")
        with jsonl.open("r", encoding="utf-8") as f:
            for line in f:
                if not line.strip():
                    continue
                obj = json.loads(line)
                if not obj.get("has_text"):
                    continue
                obj["_source_label"] = source_label
                obj["_source_dataset"] = SOURCE_DATASET_LABEL
                obj["_run_name"] = run_name
                obj["_card_key"] = make_card_key(source_label, obj)
                cards.append(obj)
    return cards


def parse_json_content(text: str) -> Dict[str, Any]:
    text = text.strip()
    if text.startswith("```"):
        text = re.sub(r"^```(?:json)?\s*", "", text)
        text = re.sub(r"\s*```$", "", text)
    return json.loads(text)


def call_deepseek(api_key: str, user_payload: List[Dict[str, Any]], model: str, timeout: int = 120) -> Dict[str, Any]:
    cards_input = []
    for c in user_payload:
        cards_input.append(
            {
                "card_key": c["_card_key"],
                "line_texts": c.get("line_texts") or [],
                "rule_extracted_hint": c.get("extracted") or {},
            }
        )
    user_msg = json.dumps({"cards": cards_input}, ensure_ascii=False)
    body = {
        "model": model,
        "messages": [
            {"role": "system", "content": SYSTEM_PROMPT},
            {"role": "user", "content": user_msg},
        ],
        "stream": False,
        "temperature": 0.1,
    }
    req = urllib.request.Request(
        DEEPSEEK_API_URL,
        data=json.dumps(body).encode("utf-8"),
        headers={"Content-Type": "application/json", "Authorization": f"Bearer {api_key}"},
        method="POST",
    )
    with urllib.request.urlopen(req, timeout=timeout) as resp:
        data = json.loads(resp.read().decode("utf-8"))
    content = data["choices"][0]["message"]["content"]
    parsed = parse_json_content(content)
    return {"parsed": parsed, "usage": data.get("usage") or {}}


def is_dirty_field(val: Any) -> bool:
    if val is None:
        return True
    s = str(val).strip()
    if not s or len(s) < 2:
        return True
    if re.fullmatch(r"[\W_]+", s):
        return True
    if re.search(r"(lyzd\d|\.com|http|www\.|@@@)", s, re.I):
        return True
    return False


def clean_field(val: Any) -> str:
    if is_dirty_field(val):
        return ""
    return str(val).strip()


def post_process_llm_item(item: Dict[str, Any]) -> Dict[str, Any]:
    mobiles_raw = item.get("mobiles") or []
    if isinstance(mobiles_raw, str):
        mobiles_raw = [mobiles_raw]
    phone_str = ";".join(str(x) for x in mobiles_raw)
    mobiles = normalize_phones(phone_str)
    return {
        "name_std": clean_field(item.get("name")),
        "title_std": clean_field(item.get("title")),
        "company_std": clean_field(item.get("company")),
        "address_std": clean_field(item.get("address")),
        "mobiles": mobiles,
        "is_key_person": bool(item.get("is_key_person")),
        "kp_role": clean_field(item.get("kp_role")),
        "card_key": item.get("card_key", ""),
    }


def load_checkpoint(cache_dir: Path, model: str) -> Dict[str, Any]:
    ckpt_path = cache_dir / "checkpoint.json"
    if ckpt_path.exists():
        state = json.loads(ckpt_path.read_text(encoding="utf-8"))
        if state.get("model") != model:
            print(f"模型切换: {state.get('model')} → {model}，已完成 {len(state.get('completed_keys', []))} 张保留续传")
            state["model"] = model
        return state
    return {
        "version": 1,
        "model": model,
        "completed_keys": [],
        "failed_batches": [],
        "status": "running",
        "started_at": utc_now_iso(),
        "updated_at": utc_now_iso(),
        "total_tokens": 0,
    }


def save_checkpoint(cache_dir: Path, state: Dict[str, Any]) -> None:
    state["updated_at"] = utc_now_iso()
    atomic_write_json(cache_dir / "checkpoint.json", state)


def rewrite_jsonl(path: Path, index: Dict[str, Dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as f:
        for key in sorted(index.keys()):
            f.write(json.dumps(index[key], ensure_ascii=False) + "\n")
        f.flush()
        os.fsync(f.fileno())


def append_jsonl(path: Path, obj: Dict[str, Any]) -> None:
    """Deprecated: use rewrite_jsonl after batch for dedup safety."""
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("a", encoding="utf-8") as f:
        f.write(json.dumps(obj, ensure_ascii=False) + "\n")
        f.flush()
        os.fsync(f.fileno())


def load_results_index(results_path: Path) -> Dict[str, Dict[str, Any]]:
    idx: Dict[str, Dict[str, Any]] = {}
    if not results_path.exists():
        return idx
    with results_path.open("r", encoding="utf-8") as f:
        for line in f:
            if line.strip():
                obj = json.loads(line)
                idx[obj["card_key"]] = obj
    return idx


def process_single_batch(
    batch: List[Dict[str, Any]],
    api_key: str,
    model: str,
    retries: int,
) -> Tuple[List[Dict[str, Any]], int]:
    """Call DeepSeek for one batch; return (records, tokens_used)."""
    last_err = None
    for attempt in range(retries + 1):
        try:
            resp = call_deepseek(api_key, batch, model)
            parsed_cards = resp["parsed"].get("cards") or []
            by_key = {x.get("card_key"): x for x in parsed_cards}
            records: List[Dict[str, Any]] = []
            for src in batch:
                ck = src["_card_key"]
                llm_item = by_key.get(ck) or {
                    "card_key": ck,
                    "name": "",
                    "title": "",
                    "company": "",
                    "address": "",
                    "mobiles": [],
                    "is_key_person": False,
                    "kp_role": "",
                }
                processed = post_process_llm_item(llm_item)
                records.append(
                    {
                        "card_key": ck,
                        "source_label": src["_source_label"],
                        "source_dataset": src["_source_dataset"],
                        "file": src.get("file"),
                        "page": src.get("page"),
                        "card_no": src.get("card_no"),
                        "grid_pos": src.get("grid_pos"),
                        "line_texts": src.get("line_texts"),
                        "rule_extracted": src.get("extracted"),
                        "llm_model": model,
                        **processed,
                    }
                )
            tokens = int((resp.get("usage") or {}).get("total_tokens") or 0)
            return records, tokens
        except (urllib.error.HTTPError, urllib.error.URLError, json.JSONDecodeError, KeyError) as e:
            last_err = str(e)
            if attempt < retries:
                time.sleep(2 * (attempt + 1))
    raise RuntimeError(f"DeepSeek 批次失败: {last_err}")


def process_batches(
    cards: List[Dict[str, Any]],
    api_key: str,
    cache_dir: Path,
    batch_size: int,
    resume: bool,
    retries: int,
    workers: int = 1,
    model: str = DEFAULT_MODEL,
    heartbeat_sec: float = DEFAULT_HEARTBEAT_SEC,
    quiet: bool = False,
) -> Dict[str, Dict[str, Any]]:
    cache_dir.mkdir(parents=True, exist_ok=True)
    results_path = cache_dir / "llm_results.jsonl"
    state = load_checkpoint(cache_dir, model)
    done: Set[str] = set(state.get("completed_keys", [])) if resume else set()
    index = load_results_index(results_path) if resume else {}
    lock = threading.Lock()

    pending = [c for c in cards if c["_card_key"] not in done]
    batches = [pending[i : i + batch_size] for i in range(0, len(pending), batch_size)]
    print(
        f"待 DeepSeek 处理: {len(pending)} / {len(cards)} 张名片"
        f"（batch={batch_size}, workers={workers}, model={model}, 批次数={len(batches)}）",
        flush=True,
    )

    reporter = ProgressReporter(
        total=len(cards),
        model=model,
        initial_done=len(done),
        initial_tokens=int(state.get("total_tokens", 0)),
        heartbeat_sec=0 if quiet else heartbeat_sec,
        enabled=not quiet,
    )

    if pending:
        state["status"] = "running"
        save_checkpoint(cache_dir, state)
        reporter.start(len(batches))
    elif resume and done:
        print("全部名片已处理，跳过 API 调用", flush=True)
        state["status"] = "completed"
        save_checkpoint(cache_dir, state)
        return index

    completed_batches = 0

    def on_batch_done(records: List[Dict[str, Any]], tokens: int) -> None:
        nonlocal completed_batches
        with lock:
            for record in records:
                index[record["card_key"]] = record
                done.add(record["card_key"])
            rewrite_jsonl(results_path, index)
            state["completed_keys"] = sorted(done)
            state["total_tokens"] = state.get("total_tokens", 0) + tokens
            state["status"] = "running"
            save_checkpoint(cache_dir, state)
            completed_batches += 1
            reporter.batch_finished(
                done=len(done),
                tokens=int(state["total_tokens"]),
                status=state["status"],
                completed_batches=completed_batches,
            )

    try:
        if workers <= 1:
            for batch in batches:
                reporter.batch_started()
                records, tokens = process_single_batch(batch, api_key, model, retries)
                on_batch_done(records, tokens)
                time.sleep(0.1)
        else:
            with ThreadPoolExecutor(max_workers=workers) as pool:
                futures = {}
                for i, b in enumerate(batches, start=1):
                    reporter.batch_started()
                    futures[pool.submit(process_single_batch, b, api_key, model, retries)] = (i, b)
                for fut in as_completed(futures):
                    batch_idx, batch = futures[fut]
                    try:
                        records, tokens = fut.result()
                        on_batch_done(records, tokens)
                    except Exception as e:
                        with lock:
                            state.setdefault("failed_batches", []).append(
                                {"batch": batch_idx, "error": str(e), "keys": [c["_card_key"] for c in batch]}
                            )
                            state["status"] = "failed"
                            save_checkpoint(cache_dir, state)
                        reporter.stop(final_status="failed")
                        raise
    finally:
        final_status = "completed" if len(done) >= len(cards) else state.get("status", "running")
        state["status"] = final_status
        save_checkpoint(cache_dir, state)
        reporter.stop(final_status=final_status)

    return index


def build_phone_rows(index: Dict[str, Dict[str, Any]]) -> List[Dict[str, Any]]:
    """Export all processed cards; annotate KP via is_key_person (do not filter)."""
    rows: List[Dict[str, Any]] = []
    for rec in index.values():
        is_kp = bool(rec.get("is_key_person"))
        mobiles = rec.get("mobiles") or []
        base = {
            SOURCE_COL: rec.get("source_label"),
            DATASET_COL: rec.get("source_dataset"),
            "file": rec.get("file"),
            "page": rec.get("page"),
            "card_no": rec.get("card_no"),
            "grid_pos": rec.get("grid_pos"),
            "name_std": rec.get("name_std"),
            "title_std": rec.get("title_std"),
            "company_std": rec.get("company_std"),
            "address_std": rec.get("address_std"),
            "kp_role": rec.get("kp_role"),
            "is_key_person": is_kp,
            "pipeline": model_pipeline_tag(rec.get("llm_model") or "deepseek_v4_flash"),
        }
        if mobiles:
            for mobile in mobiles:
                row = dict(base)
                row["phone_normalized"] = mobile
                rows.append(row)
        else:
            row = dict(base)
            row["phone_normalized"] = ""
            rows.append(row)
    return rows


def build_kp_phone_rows(index: Dict[str, Dict[str, Any]]) -> List[Dict[str, Any]]:
    """Backward-compatible alias."""
    return build_phone_rows(index)


def write_dsv4p_xlsx(output_path: Path, rows: List[Dict[str, Any]], stats: Dict[str, Any]) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "phone_processed"
    headers = [
        SOURCE_COL,
        DATASET_COL,
        "file",
        "page",
        "card_no",
        "grid_pos",
        "name_std",
        "title_std",
        "company_std",
        "address_std",
        "phone_normalized",
        "kp_role",
        "is_key_person",
        "pipeline",
    ]
    ws.append(headers)
    for row in rows:
        ws.append([row.get(h, "") for h in headers])

    ws2 = wb.create_sheet("处理统计")
    for k, v in stats.items():
        ws2.append([k, v])

    wb.save(output_path)


def compare_with_rule_pipeline(
    rule_xlsx: Path, dsv4p_rows: List[Dict[str, Any]], llm_index: Dict[str, Dict[str, Any]]
) -> Dict[str, Any]:
    rule_rows = []
    if rule_xlsx.exists():
        wb = load_workbook(rule_xlsx, read_only=True, data_only=True)
        ws = wb["phone_processed"]
        hdr = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        for row in ws.iter_rows(min_row=2, values_only=True):
            rule_rows.append({hdr[i]: row[i] for i in range(len(hdr))})

    rule_phones = {str(r.get("phone_normalized", "")).strip() for r in rule_rows if r.get("phone_normalized")}
    dsv4p_phones = {str(r.get("phone_normalized", "")).strip() for r in dsv4p_rows if r.get("phone_normalized")}

    llm_kp_cards = sum(1 for r in llm_index.values() if r.get("is_key_person"))
    llm_non_kp_cards = len(llm_index) - llm_kp_cards
    llm_with_mobile = sum(1 for r in llm_index.values() if r.get("mobiles"))
    dsv4p_kp_rows = sum(1 for r in dsv4p_rows if r.get("is_key_person"))
    dsv4p_non_kp_rows = len(dsv4p_rows) - dsv4p_kp_rows

    return {
        "rule_pipeline_rows": len(rule_rows),
        "rule_unique_phones": len(rule_phones),
        "deepseek_total_rows": len(dsv4p_rows),
        "deepseek_kp_rows": dsv4p_kp_rows,
        "deepseek_non_kp_rows": dsv4p_non_kp_rows,
        "deepseek_unique_phones": len(dsv4p_phones),
        "phones_in_both": len(rule_phones & dsv4p_phones),
        "phones_only_in_rule": len(rule_phones - dsv4p_phones),
        "phones_only_in_deepseek": len(dsv4p_phones - rule_phones),
        "llm_cards_processed": len(llm_index),
        "llm_kp_cards": llm_kp_cards,
        "llm_non_kp_cards": llm_non_kp_cards,
        "llm_kp_rate_pct": round(llm_kp_cards / len(llm_index) * 100, 1) if llm_index else 0,
        "llm_cards_with_mobile": llm_with_mobile,
    }


def export_from_cache(
    cache_dir: Path, output_xlsx: Path, rule_xlsx: Path, model: str
) -> Tuple[List[Dict[str, Any]], Dict[str, Any]]:
    index = load_results_index(cache_dir / "llm_results.jsonl")
    rows = build_phone_rows(index)
    cmp_stats = compare_with_rule_pipeline(rule_xlsx, rows, index)
    stats = {**cmp_stats, "model": model, "output_file": str(output_xlsx)}
    write_dsv4p_xlsx(output_xlsx, rows, stats)
    return rows, stats


def main() -> int:
    parser = argparse.ArgumentParser(description="DeepSeek v4-pro KP 名片后处理（基于已有 OCR）")
    parser.add_argument("--resume", action="store_true", help="断点续传")
    parser.add_argument("--export-only", action="store_true", help="仅从缓存重新导出 xlsx（不调用 API）")
    parser.add_argument("--batch-size", type=int, default=DEFAULT_BATCH_SIZE)
    parser.add_argument("--partition", default="", help="仅处理指定分区，如 2期-PartA")
    parser.add_argument("--limit", type=int, default=0, help="最多处理 N 张名片（调试）")
    parser.add_argument("--retries", type=int, default=2)
    parser.add_argument("--workers", type=int, default=DEFAULT_WORKERS, help="并行 API 批次数，默认 6")
    parser.add_argument("--model", default="", help="DeepSeek 模型，默认读 DEEPSEEK_MODEL 或 deepseek-v4-flash")
    parser.add_argument("--heartbeat", type=float, default=DEFAULT_HEARTBEAT_SEC, help="心跳打印间隔秒数，默认 15")
    parser.add_argument("--quiet", action="store_true", help="关闭终端实时进度输出")
    args = parser.parse_args()

    configure_stdio_unbuffered()
    root = project_root()
    bootstrap_dotenv(root)
    model = resolve_model(args.model)
    cache_dir = deepseek_cache_dir(root)
    output_xlsx = default_output_xlsx(root, model)
    rule_xlsx = deliverables_dir(root) / "广交会139届_名片手机号汇总.xlsx"

    if args.export_only:
        rows, cmp_stats = export_from_cache(cache_dir, output_xlsx, rule_xlsx, model)
        print(f"已从缓存导出 {len(rows)} 行（KP: {cmp_stats['deepseek_kp_rows']}, 非KP: {cmp_stats['deepseek_non_kp_rows']}）")
        print(f"输出: {output_xlsx}")
        return 0

    api_key = os.getenv("DEEPSEEK_API_KEY", "").strip()
    if not api_key:
        print("缺少 DEEPSEEK_API_KEY")
        return 2

    cards = load_cards_from_partitions(root, partition_filter=args.partition)
    if args.limit > 0:
        cards = cards[: args.limit]

    index = process_batches(
        cards=cards,
        api_key=api_key,
        cache_dir=cache_dir,
        batch_size=args.batch_size,
        resume=args.resume,
        retries=args.retries,
        workers=args.workers,
        model=model,
        heartbeat_sec=args.heartbeat,
        quiet=args.quiet,
    )
    rows, cmp_stats = export_from_cache(cache_dir, output_xlsx, rule_xlsx, model)

    ckpt = load_checkpoint(cache_dir, model)
    print(f"\n模型: {model}")
    print(f"\nDeepSeek 已处理名片: {len(index)} / {len(cards)}")
    print(f"DeepSeek 汇总行: {len(rows)}（KP: {cmp_stats['deepseek_kp_rows']}, 非KP: {cmp_stats['deepseek_non_kp_rows']}）")
    print(f"累计 tokens: {ckpt.get('total_tokens', 0)}")
    print(f"对比: 规则链路 {cmp_stats['rule_pipeline_rows']} 行 → DeepSeek {cmp_stats['deepseek_total_rows']} 行")
    print(f"手机号交集: {cmp_stats['phones_in_both']}, 仅规则: {cmp_stats['phones_only_in_rule']}, 仅DeepSeek: {cmp_stats['phones_only_in_deepseek']}")
    print(f"输出: {output_xlsx}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
