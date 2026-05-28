#!/usr/bin/env python3
"""
Compare pipeline A (rule) vs pipeline B (DeepSeek) deliverables.

Outputs: data/deliverables/广交会139届_AB链路对比报告.xlsx

Usage:
  cd src && python compare_ab_pipeline.py
"""

from __future__ import annotations

import argparse
from collections import defaultdict
from pathlib import Path
from typing import Any, Dict, List, Set, Tuple

from openpyxl import Workbook, load_workbook

from partitions import SOURCE_COL
from paths import deliverables_dir, project_root


def card_key(row: Dict[str, Any]) -> str:
    return "|".join(
        str(row.get(k, "") or "")
        for k in (SOURCE_COL, "file", "page", "card_no")
    )


def load_phone_sheet(path: Path) -> List[Dict[str, Any]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb["phone_processed"]
    hdr = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    rows: List[Dict[str, Any]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        rows.append({hdr[i]: row[i] for i in range(len(hdr))})
    wb.close()
    return rows


def aggregate_by_card(rows: List[Dict[str, Any]], phone_field: str = "phone_normalized") -> Dict[str, Dict[str, Any]]:
    by_card: Dict[str, Dict[str, Any]] = {}
    for r in rows:
        ck = card_key(r)
        if ck not in by_card:
            by_card[ck] = {
                "source_label": r.get(SOURCE_COL),
                "file": r.get("file"),
                "page": r.get("page"),
                "card_no": r.get("card_no"),
                "phones": set(),
                "name": "",
                "company": "",
                "title": "",
                "name_std": "",
                "company_std": "",
                "title_std": "",
                "is_key_person": False,
                "kp_role": "",
            }
        phone = str(r.get(phone_field, "") or "").strip()
        if phone:
            by_card[ck]["phones"].add(phone)
        for field in ("name", "company", "title", "name_std", "company_std", "title_std", "kp_role"):
            val = str(r.get(field, "") or "").strip()
            if val and not by_card[ck].get(field):
                by_card[ck][field] = val
        if r.get("is_key_person"):
            by_card[ck]["is_key_person"] = True
    return by_card


def build_comparison(
    rule_rows: List[Dict[str, Any]], ds_rows: List[Dict[str, Any]]
) -> Tuple[Dict[str, Any], List[List[Any]], List[List[Any]], List[List[Any]]]:
    rule_cards = aggregate_by_card(rule_rows)
    ds_cards = aggregate_by_card(ds_rows)

    rule_phones: Set[str] = set()
    ds_phones: Set[str] = set()
    for c in rule_cards.values():
        rule_phones |= c["phones"]
    for c in ds_cards.values():
        ds_phones |= c["phones"]

    all_keys = set(rule_cards) | set(ds_cards)
    phone_match = phone_a_only = phone_b_only = phone_diff = 0
    name_filled_a = name_filled_b = 0
    company_filled_a = company_filled_b = 0

    diff_rows: List[List[Any]] = []
    for ck in sorted(all_keys):
        ra = rule_cards.get(ck)
        rb = ds_cards.get(ck)
        pa = ra["phones"] if ra else set()
        pb = rb["phones"] if rb else set()
        if pa == pb and pa:
            phone_match += 1
        elif pa and pb and pa != pb:
            phone_diff += 1
            diff_rows.append(
                [
                    ck,
                    ra.get("source_label") if ra else rb.get("source_label"),
                    ";".join(sorted(pa)),
                    ";".join(sorted(pb)),
                    "手机号集合不一致",
                ]
            )
        if pa and not pb:
            phone_a_only += 1
            if len(diff_rows) < 500:
                diff_rows.append([ck, ra.get("source_label"), ";".join(sorted(pa)), "", "仅链路A有号"])
        if pb and not pa:
            phone_b_only += 1
            if len(diff_rows) < 500:
                diff_rows.append([ck, rb.get("source_label"), "", ";".join(sorted(pb)), "仅链路B有号"])

        if ra and str(ra.get("name", "")).strip():
            name_filled_a += 1
        if rb and str(rb.get("name_std", "")).strip():
            name_filled_b += 1
        if ra and str(ra.get("company", "")).strip():
            company_filled_a += 1
        if rb and str(rb.get("company_std", "")).strip():
            company_filled_b += 1

    kp_cards = sum(1 for c in ds_cards.values() if c["is_key_person"])
    ds_rows_kp = sum(1 for r in ds_rows if r.get("is_key_person"))
    ds_rows_non_kp = len(ds_rows) - ds_rows_kp

    overview = {
        "rule_phone_rows": len(rule_rows),
        "rule_unique_phones": len(rule_phones),
        "rule_cards_with_phone": sum(1 for c in rule_cards.values() if c["phones"]),
        "deepseek_phone_rows": len(ds_rows),
        "deepseek_unique_phones": len(ds_phones),
        "deepseek_cards_in_export": len(ds_cards),
        "deepseek_kp_rows": ds_rows_kp,
        "deepseek_non_kp_rows": ds_rows_non_kp,
        "deepseek_kp_cards": kp_cards,
        "phones_in_both": len(rule_phones & ds_phones),
        "phones_only_in_rule": len(rule_phones - ds_phones),
        "phones_only_in_deepseek": len(ds_phones - rule_phones),
        "cards_phone_set_equal": phone_match,
        "cards_phone_set_diff": phone_diff,
        "cards_only_rule_has_phone": phone_a_only,
        "cards_only_deepseek_has_phone": phone_b_only,
        "cards_with_name_rule": name_filled_a,
        "cards_with_name_std_deepseek": name_filled_b,
        "cards_with_company_rule": company_filled_a,
        "cards_with_company_std_deepseek": company_filled_b,
    }

    only_rule_phones = sorted(rule_phones - ds_phones)[:2000]
    only_ds_phones = sorted(ds_phones - rule_phones)[:2000]
    phone_detail: List[List[Any]] = [
        ["仅链路A手机号", p] for p in only_rule_phones
    ] + [["仅链路B手机号", p] for p in only_ds_phones]

    return overview, diff_rows, phone_detail, list(overview.items())


def recommendation_text(overview: Dict[str, Any]) -> List[List[Any]]:
    both = overview["phones_in_both"]
    only_a = overview["phones_only_in_rule"]
    only_b = overview["phones_only_in_deepseek"]
    kp = overview["deepseek_kp_cards"]

    lines = [
        ["维度", "链路 A（规则）", "链路 B（DeepSeek flash）", "建议"],
        [
            "手机号召回（唯一号）",
            f"{overview['rule_unique_phones']} 个",
            f"{overview['deepseek_unique_phones']} 个",
            f"A 唯一号略多；但有 {overview['phones_only_in_deepseek']} 个号仅 B 识别、{overview['phones_only_in_rule']} 个仅 A，需按场景合并",
        ],
        [
            "手机号精确对齐",
            f"交集 {both}",
            f"交集 {both}",
            f"约 {only_a} 个仅 A、{only_b} 个仅 B，需抽样核对 OCR 噪声 vs LLM 纠错",
        ],
        [
            "姓名/公司/职位",
            "关键词首行启发式，无 LLM",
            "LLM 标准化 *_std + KP 标注",
            "要联系人质量、职级筛选 → 优先 B；要低成本全量手机号 → A 足够",
        ],
        [
            "KP 决策人",
            "无",
            f"{kp} 张名片 is_key_person=true",
            "外呼/BD 筛决策层用 B 的 is_key_person；全量触达用手机号 sheet 用 A 或 B 均可",
        ],
        [
            "成本与稳定性",
            "仅火山 OCR，本地规则",
            "9700 次 LLM，约 886 万 tokens",
            "A 为基线交付；B 为增值层，不替代 OCR",
        ],
        [
            "综合结论",
            "—",
            "—",
            "最终手机号以 A 为保守基线；B 用于字段标准化 + KP 分层 + 补召回。推荐对外交付：A 全量号 + B 的 KP 子集或合并去重后人工抽检",
        ],
    ]
    return lines


def write_report(
    output: Path,
    overview: Dict[str, Any],
    diff_rows: List[List[Any]],
    phone_detail: List[List[Any]],
    conclusion: List[List[Any]],
) -> None:
    wb = Workbook()
    ws0 = wb.active
    ws0.title = "概览"
    ws0.append(["指标", "值"])
    for k, v in overview.items():
        ws0.append([k, v])

    ws1 = wb.create_sheet("结论与选型")
    for row in conclusion:
        ws1.append(row)

    ws2 = wb.create_sheet("名片级差异样例")
    ws2.append(["card_key", "source_label", "phones_A", "phones_B", "备注"])
    for row in diff_rows[:500]:
        ws2.append(row)

    ws3 = wb.create_sheet("手机号单边样例")
    ws3.append(["类型", "phone_normalized"])
    for row in phone_detail[:4000]:
        ws3.append(row)

    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)


def main() -> int:
    parser = argparse.ArgumentParser(description="对比链路 A/B 汇总结果")
    root = project_root()
    dlv = deliverables_dir(root)
    parser.add_argument(
        "--rule-xlsx",
        type=Path,
        default=dlv / "广交会139届_名片手机号汇总.xlsx",
    )
    parser.add_argument(
        "--deepseek-xlsx",
        type=Path,
        default=dlv / "广交会139届_名片手机号汇总_dsv4flash.xlsx",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=dlv / "广交会139届_AB链路对比报告.xlsx",
    )
    args = parser.parse_args()

    if not args.rule_xlsx.exists():
        print(f"缺少链路 A 文件: {args.rule_xlsx}")
        return 2
    if not args.deepseek_xlsx.exists():
        print(f"缺少链路 B 文件: {args.deepseek_xlsx}")
        return 2

    rule_rows = load_phone_sheet(args.rule_xlsx)
    ds_rows = load_phone_sheet(args.deepseek_xlsx)
    overview, diff_rows, phone_detail, _ = build_comparison(rule_rows, ds_rows)
    conclusion = recommendation_text(overview)
    write_report(args.output, overview, diff_rows, phone_detail, conclusion)

    print("链路 A:", args.rule_xlsx.name, f"→ {len(rule_rows)} 行, {overview['rule_unique_phones']} 唯一号")
    print("链路 B:", args.deepseek_xlsx.name, f"→ {len(ds_rows)} 行, {overview['deepseek_unique_phones']} 唯一号")
    print(f"交集 {overview['phones_in_both']} | 仅A {overview['phones_only_in_rule']} | 仅B {overview['phones_only_in_deepseek']}")
    print(f"报告: {args.output}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
